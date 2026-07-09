from __future__ import annotations

import math
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import Side
import pandas as pd

# -----------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_ROOT = PROJECT_ROOT / "Output_data"
PAINT_DIR = PROJECT_ROOT / "Paint"

MODEL = "PFSDF"
EDP = "PFA"
REFERENCE_TEMP = 20.0
TEMPERATURES = [-20.0, -10.0, 0.0, 10.0, 20.0, 30.0, 40.0]

TEMPERATURE_WEIGHT_FILE = PAINT_DIR / f"regional_temperature_weights_{MODEL}.csv"

CITY_CONFIGS = [
    {"city": "Harbin", "label": "Harbin", "color": "#2a7fde", "linestyle": "-"},
    {"city": "Yinchuan", "label": "Yinchuan", "color": "#8abf4b", "linestyle": "-"},
    {"city": "Chongqing", "label": "Chongqing", "color": "#f0a03a", "linestyle": "-"},
    {"city": "20C", "label": "20掳C", "color": "black", "linestyle": (0, (5, 3))},
]

BASE_ANNOTATION_OFFSETS = {
    "Harbin": (-34, -14),
    "Yinchuan": (-16, 12),
    "Chongqing": (12, -8),
    "20C": (28, 10),
}

# EDP 缁熶竴閰嶇疆
# 3. DS 闃堝€间笌榛樿鏀惧ぇ瀹藉害
# 4. 灞€閮ㄥ浘鎵嬪姩鍧愭爣鑼冨洿
EDP_CONFIGS = {
    "IDR": {
        "plot_scale": 100.0,
        "plot_xlabel": "IDR (%)",
        "plot_unit": "%",
        "overview_figsize": (8, 6),
        "zoom_figsize": (8, 6),
        "overview_xlim": (0.0, 10.0),
        "overview_ylim": (1e-6, 1.0),
        "overview_xticks": np.arange(0.0, 10.1, 2.0),
        "ds_configs": [
            {"name": "DS-1", "threshold": 0.02, "zoom_half_width": 0.50},
            {"name": "DS-2", "threshold": 0.03, "zoom_half_width": 0.50},
            {"name": "DS-3", "threshold": 0.05, "zoom_half_width": 0.70},
        ],
        "zoom_axis_configs": {
            "DS-1": {"xlim": None, "ylim": [4e-3, 1.4e-2], "xticks": [1.5, 2.0, 2.5], "yticks": np.linspace(0.004, 0.014, 4)},
            "DS-2": {"xlim": None, "ylim": [2e-3, 5e-3], "xticks": [2.5, 3.0, 3.5], "yticks": np.linspace(0.002, 0.005, 4)},
            "DS-3": {"xlim": [4.5, 5.5], "ylim": [4e-4, 1.2e-3], "xticks": [4.5, 5.0, 5.5], "yticks": np.linspace(0.0004, 0.0012, 4)},
        },
        "annotation_offsets": {
            "DS-1": {"Harbin": (2, 30), "Yinchuan": (2, 15), "Chongqing": (2, -65), "20C": (2, -80)},
            "DS-2": {"Harbin": (2, 50), "Yinchuan": (2, 35), "Chongqing": (2, 25), "20C": (2, -60)},
            "DS-3": {"Harbin": (2, 30), "Yinchuan": (2, 20), "Chongqing": (2, -70), "20C": (2, -80)},
        },
    },
    "RIDR": {
        "plot_scale": 100.0,
        "plot_xlabel": "RIDR (%)",
        "plot_unit": "%",
        "overview_figsize": (8, 6),
        "zoom_figsize": (8, 6),
        "overview_xlim": (0.0, 2.0),
        "overview_ylim": (1e-4, 1.0),
        "overview_xticks": np.arange(0.0, 2.01, 0.5),
        "ds_configs": [
            {"name": "DS-1", "threshold": 0.002, "zoom_half_width": 0.10},
            {"name": "DS-2", "threshold": 0.005, "zoom_half_width": 0.10},
            {"name": "DS-3", "threshold": 0.010, "zoom_half_width": 0.10},
        ],
        "zoom_axis_configs": {
            "DS-1": {"xlim": None, "ylim": [2e-3, 8e-3], "xticks": None, "yticks": np.linspace(0.002, 0.008, 4)},
            "DS-2": {"xlim": None, "ylim": [1e-3, 2.5e-3], "xticks": None, "yticks": np.linspace(0.001, 0.0025, 4)},
            "DS-3": {"xlim": None, "ylim": [7e-4, 12e-4], "xticks": None, "yticks": np.linspace(0.0007, 0.0012, 4)},
        },
        "annotation_offsets": {
            "DS-1": {"Harbin": (2, 30), "Yinchuan": (2, 15), "Chongqing": (2, -65), "20C": (2, -80)},
            "DS-2": {"Harbin": (2, 45), "Yinchuan": (2, 35), "Chongqing": (2, -55), "20C": (2, -65)},
            "DS-3": {"Harbin": (2, 30), "Yinchuan": (2, 28), "Chongqing": (2, -70), "20C": (2, -80)},
        },
    },
    "PFA": {
        "plot_scale": 1.0,
        "plot_xlabel": "PFA (g)",
        "plot_unit": "g",
        "overview_figsize": (8, 6),
        "zoom_figsize": (8, 6),
        "overview_xlim": (0.0, 2),
        "overview_ylim": (1e-3, 1.0),
        "overview_xticks": np.arange(0.0, 2.1, 0.5),
        "ds_configs": [
            {"name": "DS-1", "threshold": 0.5, "zoom_half_width": 0.10},
            {"name": "DS-2", "threshold": 1.0, "zoom_half_width": 0.10},
            {"name": "DS-3", "threshold": 1.5, "zoom_half_width": 0.12},
        ],
        "zoom_axis_configs": {
            "DS-1": {"xlim": None, "ylim": [2e-2, 7e-2], "xticks": [0.4,0.5,0.6], "yticks": np.linspace(0.02, 0.07, 4)},
            "DS-2": {"xlim": None, "ylim": [6e-3, 1.6e-2], "xticks": [0.9,1.0,1.1], "yticks": np.linspace(0.006, 0.016, 4)},
            "DS-3": {"xlim": [1.4,1.6], "ylim": [3e-3, 6.5e-3], "xticks": [1.4,1.5,1.6], "yticks": np.linspace(0.003, 0.0065, 4)},
        },
        "annotation_offsets": {
            "DS-1": {"Harbin": (2, 45), "Yinchuan": (2, 15), "Chongqing": (2, -65), "20C": (2, -100)},
            "DS-2": {"Harbin": (2, 50), "Yinchuan": (2, 15), "Chongqing": (2, -60), "20C": (2, -80)},
            "DS-3": {"Harbin": (2, 50), "Yinchuan": (2, 15), "Chongqing": (2, -70), "20C": (2, -80)},
        },
    },
}

if EDP not in EDP_CONFIGS:
    raise ValueError(f"涓嶆敮鎸佺殑 EDP 绫诲瀷: {EDP}锛屽彲閫夊€间负 {list(EDP_CONFIGS)}")

ACTIVE_EDP_CONFIG = EDP_CONFIGS[EDP]
PLOT_SCALE = float(ACTIVE_EDP_CONFIG["plot_scale"])
PLOT_XLABEL = str(ACTIVE_EDP_CONFIG["plot_xlabel"])
PLOT_UNIT = str(ACTIVE_EDP_CONFIG["plot_unit"])
OVERVIEW_FIGSIZE = tuple(ACTIVE_EDP_CONFIG["overview_figsize"])
ZOOM_FIGSIZE = tuple(ACTIVE_EDP_CONFIG["zoom_figsize"])
OVERVIEW_XLIM = tuple(ACTIVE_EDP_CONFIG["overview_xlim"])
OVERVIEW_YLIM = tuple(ACTIVE_EDP_CONFIG["overview_ylim"])
OVERVIEW_XTICKS = np.asarray(ACTIVE_EDP_CONFIG["overview_xticks"], dtype=float)
DS_CONFIGS = list(ACTIVE_EDP_CONFIG["ds_configs"])
ZOOM_AXIS_CONFIGS = dict(ACTIVE_EDP_CONFIG["zoom_axis_configs"])
ANNOTATION_OFFSETS = dict(ACTIVE_EDP_CONFIG["annotation_offsets"])

OUTPUT_STEM = f"Regional_hazard_{MODEL}_{EDP}"
SHOW_FIGURE = False


# -----------------------------
# 缁樺浘鍏ㄥ眬椋庢牸
# -----------------------------
def configure_matplotlib() -> None:
    plt.rcParams["font.family"] = "serif"
    plt.rcParams["font.serif"] = ["Times New Roman", "SimSun", "STSong", "DejaVu Serif"]
    plt.rcParams["mathtext.fontset"] = "stix"
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["axes.linewidth"] = 1.0


# -----------------------------
def load_temperature_weights(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Base hazard curve file not found: {path}")

    df = pd.read_csv(path)
    required_columns = {"city", "mapped_temperature_C", "region_probability"}
    missing_columns = required_columns.difference(df.columns)
    if missing_columns:
        raise ValueError(f"{path.name} missing columns: {sorted(missing_columns)}")

    df["mapped_temperature_C"] = df["mapped_temperature_C"].astype(float)
    df["region_probability"] = df["region_probability"].astype(float)

    for city in df["city"].unique():
        city_prob_sum = df.loc[df["city"] == city, "region_probability"].sum()
        if not np.isclose(city_prob_sum, 1.0, atol=1e-8):
            raise ValueError(f"{city} 鐨勫尯鍩熸俯搴︽鐜囧拰涓嶄负 1锛屽綋鍓嶄负 {city_prob_sum:.10f}")

    return df.sort_values(["city", "mapped_temperature_C"]).reset_index(drop=True)


def hazard_curve_path(model: str, temperature: float, edp: str) -> Path:
    temp_text = str(int(temperature)) if float(temperature).is_integer() else f"{temperature:g}"
    return OUTPUT_ROOT / f"MC8_{model}_{temp_text}" / "MC8_IDA_data_frag" / f"hazard_curve_{edp}.out"


def load_hazard_curve(model: str, temperature: float, edp: str) -> pd.DataFrame:
    path = hazard_curve_path(model, temperature, edp)
    if not path.exists():
        raise FileNotFoundError(f"Hazard curve file not found for {temperature:g} C {edp}: {path}")

    data = np.loadtxt(path)
    if data.ndim != 2 or data.shape[1] != 2:
        raise ValueError(f"{path.name} EDP and lambda_EDP values must be positive.")

    df = pd.DataFrame(data, columns=["edp", "lambda_edp"]).sort_values("edp").reset_index(drop=True)
    if (df["edp"] <= 0).any() or (df["lambda_edp"] <= 0).any():
        raise ValueError(f"{path.name} EDP and lambda_EDP values must be positive.")
    return df


def load_temperature_hazard_curves(model: str, edp: str, temperatures: list[float]) -> dict[float, pd.DataFrame]:
    curves = {float(temp): load_hazard_curve(model, float(temp), edp) for temp in temperatures}
    reference_grid = curves[float(temperatures[0])]["edp"].to_numpy(dtype=float)
    for temp, curve in curves.items():
        current_grid = curve["edp"].to_numpy(dtype=float)
        if len(current_grid) != len(reference_grid) or not np.allclose(current_grid, reference_grid, atol=1e-12):
            raise ValueError(f"Hazard curve grid at {temp:g} C differs from reference grid.")
    return curves


# -----------------------------
def regional_hazard_curve(
    city: str,
    weight_table: pd.DataFrame,
    temperature_curves: dict[float, pd.DataFrame],
) -> pd.DataFrame:
    city_weights = weight_table.loc[weight_table["city"] == city, ["mapped_temperature_C", "region_probability"]]
    if city_weights.empty:
        raise ValueError(f"Unknown city in regional weights: {city}")

    base_curve = next(iter(temperature_curves.values()))
    weighted_lambda = np.zeros(len(base_curve), dtype=float)

    for row in city_weights.itertuples(index=False):
        temp = float(row.mapped_temperature_C)
        probability = float(row.region_probability)
        if temp not in temperature_curves:
            raise ValueError(f"Hazard curve grid at {temp:g} C differs from reference grid.")
        weighted_lambda += probability * temperature_curves[temp]["lambda_edp"].to_numpy(dtype=float)

    return pd.DataFrame({"edp": base_curve["edp"].to_numpy(dtype=float), "lambda_edp": weighted_lambda})


def build_curves(weight_table: pd.DataFrame, temperature_curves: dict[float, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    curves: dict[str, pd.DataFrame] = {}
    for config in CITY_CONFIGS:
        city = config["city"]
        if city == "20C":
            curves[city] = temperature_curves[REFERENCE_TEMP].copy()
        else:
            curves[city] = regional_hazard_curve(city, weight_table, temperature_curves)
    return curves


# -----------------------------
# DS 鎸囨爣鎻愬彇
# -----------------------------
def interpolate_lambda(curve: pd.DataFrame, edp_threshold: float) -> float:
    x_values = curve["edp"].to_numpy(dtype=float)
    y_values = curve["lambda_edp"].to_numpy(dtype=float)
    if not (float(x_values.min()) <= edp_threshold <= float(x_values.max())):
        raise ValueError(f"EDP threshold {edp_threshold:.4f} is outside curve range [{x_values.min():.4f}, {x_values.max():.4f}]")
    return float(np.interp(edp_threshold, x_values, y_values))


def extract_ds_results(curves: dict[str, pd.DataFrame], ds_configs: list[dict[str, float | str]]) -> pd.DataFrame:
    rows: list[dict[str, float | str]] = []
    for ds in ds_configs:
        ds_name = str(ds["name"])
        threshold = float(ds["threshold"])
        row: dict[str, float | str] = {
            "EDP": EDP,
            "DS": ds_name,
            "Threshold": threshold,
            f"Threshold ({PLOT_UNIT})": threshold * PLOT_SCALE,
        }
        for config in CITY_CONFIGS:
            city_key = config["city"]
            city_label = str(config["label"])
            lambda_edp = interpolate_lambda(curves[city_key], threshold)
            p50 = 1.0 - math.exp(-50.0 * lambda_edp)
            row[f"{city_label}_lambda_EDP"] = lambda_edp
            row[f"{city_label}_P50 (%)"] = 100.0 * p50
        rows.append(row)
    return pd.DataFrame(rows)


# -----------------------------
# 閫氱敤缁樺浘杈呭姪
# -----------------------------
def style_axes(ax: plt.Axes, log_y: bool) -> None:
    if log_y:
        ax.set_yscale("log")
    ax.grid(which="major", linestyle="-", linewidth=0.45, color="#b9b9b9", alpha=0.85)
    ax.grid(which="minor", linestyle=":", linewidth=0.35, color="#c7c7c7", alpha=0.85)
    ax.tick_params(axis="both", which="major", direction="in", top=True, right=True, labelsize=11)
    ax.tick_params(axis="both", which="minor", direction="in", top=True, right=True)


def scientific_formatter(value: float) -> str:
    if np.isclose(value, 0.0):
        return "0"
    exponent = int(np.floor(np.log10(abs(value))))
    mantissa = value / (10**exponent)
    return rf"${mantissa:.1f}\times10^{{{exponent}}}$"


def apply_scientific_yticks(ax: plt.Axes) -> None:
    ticks = ax.get_yticks()
    labels = [scientific_formatter(tick) for tick in ticks]
    ax.set_yticks(ticks)
    ax.set_yticklabels(labels)


def add_curve_lines(ax: plt.Axes, curves: dict[str, pd.DataFrame]) -> None:
    for config in CITY_CONFIGS:
        city = config["city"]
        curve = curves[city]
        ax.plot(
            curve["edp"] * PLOT_SCALE,
            curve["lambda_edp"],
            color=config["color"],
            linestyle=config["linestyle"],
            linewidth=1.5,
            label=config["label"],
        )


def save_figure(fig: plt.Figure, stem: str) -> None:
    png_path = PAINT_DIR / f"{stem}.png"
    fig.savefig(png_path, dpi=1000, bbox_inches="tight")


def resolve_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    for idx in range(1, 100):
        candidate = path.with_name(f"{path.stem}_{idx}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"鏃犳硶鎵惧埌鍙敤杈撳嚭璺緞: {path}")


# -----------------------------
# 鎬诲浘缁樺埗
# -----------------------------
def plot_overview(curves: dict[str, pd.DataFrame], ds_configs: list[dict[str, float | str]]) -> None:
    fig, ax = plt.subplots(figsize=OVERVIEW_FIGSIZE, dpi=300)
    add_curve_lines(ax, curves)
    style_axes(ax, log_y=True)

    ax.set_xlim(*OVERVIEW_XLIM)
    ax.set_xticks(OVERVIEW_XTICKS)
    ax.set_ylim(*OVERVIEW_YLIM)
    ax.set_xlabel(PLOT_XLABEL, fontsize=25)
    ax.set_ylabel(r"$\lambda_{\mathrm{EDP}}$", fontsize=25)
    ax.tick_params(direction="in", top=True, right=True, labelsize=25)

    y_text = OVERVIEW_YLIM[0] * 1.8
    for ds in ds_configs:
        threshold_plot = float(ds["threshold"]) * PLOT_SCALE
        ax.axvline(threshold_plot, color="#666666", linestyle="--", linewidth=0.9)
        ax.text(threshold_plot, y_text, str(ds["name"]), ha="center", va="bottom", fontsize=25)

    ax.legend(loc="upper right", frameon=False, fontsize=18, handlelength=2.2)
    fig.subplots_adjust(left=0.15, right=0.97, bottom=0.14, top=0.97)
    save_figure(fig, f"{OUTPUT_STEM}_overview")

    if SHOW_FIGURE:
        plt.show()
    else:
        plt.close(fig)


# -----------------------------
# 灞€閮ㄦ斁澶у浘杈呭姪
# -----------------------------
def compute_zoom_limits(curves: dict[str, pd.DataFrame], center_plot: float, half_width_plot: float) -> tuple[float, float, float, float]:
    x_min = max(0.0, center_plot - half_width_plot)
    x_max = center_plot + half_width_plot
    y_min = float("inf")
    y_max = float("-inf")

    for curve in curves.values():
        x_plot = curve["edp"].to_numpy(dtype=float) * PLOT_SCALE
        y = curve["lambda_edp"].to_numpy(dtype=float)
        mask = (x_plot >= x_min) & (x_plot <= x_max)
        if not np.any(mask):
            continue
        y_min = min(y_min, float(np.min(y[mask])))
        y_max = max(y_max, float(np.max(y[mask])))

    if not np.isfinite(y_min) or not np.isfinite(y_max):
        raise ValueError(f"Cannot determine zoom range near {center_plot:.3f}.")

    span = y_max - y_min
    pad = 0.12 * span if span > 0 else 0.05 * y_max
    return x_min, x_max, y_min - pad, y_max + pad


def resolve_zoom_axis_config(
    ds_name: str,
    curves: dict[str, pd.DataFrame],
    center_plot: float,
    half_width_plot: float,
) -> tuple[tuple[float, float], tuple[float, float], np.ndarray | None, np.ndarray | None]:
    config = ZOOM_AXIS_CONFIGS.get(ds_name, {})
    auto_x_min, auto_x_max, auto_y_min, auto_y_max = compute_zoom_limits(curves, center_plot, half_width_plot)

    xlim = tuple(config["xlim"]) if config.get("xlim") is not None else (auto_x_min, auto_x_max)
    ylim = tuple(config["ylim"]) if config.get("ylim") is not None else (auto_y_min, auto_y_max)
    xticks = None if config.get("xticks") is None else np.asarray(config["xticks"], dtype=float)
    yticks = None if config.get("yticks") is None else np.asarray(config["yticks"], dtype=float)
    return xlim, ylim, xticks, yticks


def annotation_offsets(ds_name: str) -> dict[str, tuple[int, int]]:
    return ANNOTATION_OFFSETS.get(ds_name, BASE_ANNOTATION_OFFSETS)


def annotate_zoom_values(ax: plt.Axes, curves: dict[str, pd.DataFrame], ds_name: str, threshold: float) -> None:
    x_plot = threshold * PLOT_SCALE
    offsets = annotation_offsets(ds_name)

    for config in CITY_CONFIGS:
        city = config["city"]
        lambda_edp = interpolate_lambda(curves[city], threshold)
        ax.plot(
            [x_plot],
            [lambda_edp],
            marker="o",
            markersize=3.8,
            color=config["color"],
            markerfacecolor="white" if city != "20C" else config["color"],
            markeredgewidth=0.9,
            linestyle="None",
        )
        dx, dy = offsets.get(city, BASE_ANNOTATION_OFFSETS[city])
        ax.annotate(
            f"{lambda_edp:.3e}",
            xy=(x_plot, lambda_edp),
            xytext=(dx, dy),
            textcoords="offset points",
            color=config["color"],
            fontsize=23,
        )


# -----------------------------
# DS 灞€閮ㄥ浘缁樺埗
# -----------------------------
def plot_ds_zoom(curves: dict[str, pd.DataFrame], ds_name: str, threshold: float, zoom_half_width: float) -> None:
    center_plot = threshold * PLOT_SCALE
    half_width_plot = zoom_half_width
    xlim, ylim, xticks, yticks = resolve_zoom_axis_config(ds_name, curves, center_plot, half_width_plot)
    x_min, x_max = xlim
    y_min, y_max = ylim

    fig, ax = plt.subplots(figsize=ZOOM_FIGSIZE, dpi=300)
    add_curve_lines(ax, curves)
    style_axes(ax, log_y=False)

    ax.set_xlim(*xlim)
    ax.set_ylim(*ylim)
    ax.set_xlabel(PLOT_XLABEL, fontsize=25)
    ax.set_ylabel(r"$\lambda_{\mathrm{EDP}}$", fontsize=25)
    ax.tick_params(direction="in", top=True, right=True, labelsize=25)
    ax.axvline(center_plot, color="#666666", linestyle="--", linewidth=0.9)
    ax.text(center_plot, y_max - 0.10 * (y_max - y_min), ds_name, ha="center", va="bottom", fontsize=25)

    if xticks is None:
        if PLOT_UNIT == "%":
            step = 0.1 if (x_max - x_min) <= 1.2 else 0.2
        else:
            step = 0.05 if (x_max - x_min) <= 0.6 else 0.1
        tick_start = round(x_min / step) * step
        tick_end = round(x_max / step) * step + 0.001
        ax.set_xticks(np.arange(tick_start, tick_end, step))
    else:
        ax.set_xticks(xticks)

    if yticks is not None:
        ax.set_yticks(yticks)
    apply_scientific_yticks(ax)

    annotate_zoom_values(ax, curves, ds_name, threshold)
    ax.legend(loc="lower left", frameon=False, fontsize=20, handlelength=2.0)
    fig.subplots_adjust(left=0.17, right=0.97, bottom=0.16, top=0.96)
    save_figure(fig, f"{OUTPUT_STEM}_{ds_name}")

    if SHOW_FIGURE:
        plt.show()
    else:
        plt.close(fig)


# -----------------------------
# 缁撴灉瀵煎嚭
# -----------------------------
def export_summary_table(summary_df: pd.DataFrame) -> Path:
    output_path = PAINT_DIR / f"{OUTPUT_STEM}_summary.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"

    raw_sheet = workbook.create_sheet("RawData")
    raw_sheet.append(list(summary_df.columns))
    for row in summary_df.itertuples(index=False):
        raw_sheet.append(list(row))

    thin_center = Alignment(horizontal="center", vertical="center")
    header_font = Font(name="Times New Roman", size=14, bold=False)
    body_font = Font(name="Times New Roman", size=13, bold=False)
    border_top = Border(top=Side(style="medium", color="000000"))
    border_bottom = Border(bottom=Side(style="medium", color="000000"))

    cities = ["Harbin", "Yinchuan", "Chongqing", "20掳C"]
    city_labels = {"Harbin": "Harbin", "Yinchuan": "Yinchuan", "Chongqing": "Chongqing", "20C": "20C", "20\u00b0C": "20C"}

    superscript_map = str.maketrans("-0123456789", "鈦烩伆鹿虏鲁鈦粹伒鈦垛伔鈦糕伖")

    def format_lambda_text(value: float) -> str:
        if np.isclose(value, 0.0):
            return "0"
        exponent = int(np.floor(np.log10(abs(value))))
        mantissa = value / (10**exponent)
        return f"{mantissa:.3f}脳10{str(exponent).translate(superscript_map)}"

    # 涓よ琛ㄥご
    worksheet.merge_cells("A1:B2")
    worksheet["A1"] = "DS闃舵"

    start_col = 3
    for city in cities:
        worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
        worksheet.cell(row=1, column=start_col).value = city_labels[city]
        worksheet.cell(row=2, column=start_col).value = "位EDP"
        worksheet.cell(row=2, column=start_col + 1).value = "PEDP(%)"
        start_col += 2

    data_end_row = data_start_row + len(summary_df) - 1
    worksheet.merge_cells(start_row=data_start_row, start_column=1, end_row=data_end_row, end_column=1)
    worksheet.cell(row=data_start_row, column=1).value = EDP

    for idx, (_, row) in enumerate(summary_df.iterrows(), start=data_start_row):
        worksheet.cell(row=idx, column=2).value = row["DS"]
        write_col = 3
        for city in cities:
            lambda_value = row[f"{city}_lambda_EDP"]
            p50_value = row[f"{city}_P50 (%)"]
            worksheet.cell(row=idx, column=write_col).value = format_lambda_text(float(lambda_value))
            worksheet.cell(row=idx, column=write_col + 1).value = f"{float(p50_value):.3f}"
            write_col += 2

        for cell in row:
            cell.alignment = thin_center
            cell.font = header_font if cell.row <= 2 else body_font

    for col in range(1, 11):
        worksheet.cell(row=1, column=col).border = border_top
        worksheet.cell(row=data_end_row, column=col).border = border_bottom

    column_widths = {
        "A": 10,
        "B": 10,
        "C": 14,
        "D": 12,
        "E": 14,
        "F": 12,
        "G": 14,
        "H": 12,
        "I": 14,
        "J": 12,
    }
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 22

    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = resolve_output_path(output_path)
        workbook.save(fallback_path)
        return fallback_path


# -----------------------------
def main() -> None:
    configure_matplotlib()
    weight_table = load_temperature_weights(TEMPERATURE_WEIGHT_FILE)
    temperature_curves = load_temperature_hazard_curves(MODEL, EDP, TEMPERATURES)
    curves = build_curves(weight_table, temperature_curves)
    summary_df = extract_ds_results(curves, DS_CONFIGS)

    plot_overview(curves, DS_CONFIGS)
    for ds in DS_CONFIGS:
        plot_ds_zoom(
            curves,
            str(ds["name"]),
            float(ds["threshold"]),
            float(ds["zoom_half_width"]),
        )
    summary_path = export_summary_table(summary_df)

    print(f"Generated overview and 3 zoom figures in: {PAINT_DIR}")
    print(f"Generated Excel summary: {summary_path}")


if __name__ == "__main__":
    main()
