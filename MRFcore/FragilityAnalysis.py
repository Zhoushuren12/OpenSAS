import os
import sys
from pathlib import Path
from typing import Literal
from numbers import Real
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from scipy.stats import norm
from scipy.interpolate import interp1d
import openpyxl as px
from loguru import logger
from wsection import WSection
from MRFcore.get_SSF import get_SFF
from MRFcore.get_acceptable_ACMR import get_acceptable_ACMR
from scipy.interpolate import PchipInterpolator
from scipy.signal import savgol_filter
if __name__ == "__main__":
    sys.path.append(str(Path(__file__).parent.parent))

"""
最后更新：
2024-04-07: 优化画图代码，增加保存结果图像文件
2024-05-31: 功能增强，可同时处理多种工程需求参数类型，代码结构优化
"""

logger.remove()
logger.add(
    sink=sys.stdout,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> <red>|</red> <level>{level}</level> <red>|</red> <level>{message}</level>",
    level="DEBUG"
)


def get_x(x: list, y: list, y0: float) -> float:
    """获得横线y=y0与给定曲线的交点横坐标

    Args:
        x (list): 输入曲线的横坐标序列\n
        y (list): 输入曲线的纵坐标序列\n
        y0 (float): 横直线y = y0\n

    Returns:
        float: 曲线与横线交点横坐标
    """
    if y0 < min(y):
        raise ValueError(f'【Error】y0 < min(y) ({y0} < {min(y)})')
    if y0 > max(y):
        raise ValueError(f'【Error】y0 > max(y) ({y0} > {max(y)})')
    for i in range(len(y) - 1):
        if y[i] == y0:
            x0 = x[i]
            return x0
        elif y[i] < y0 <= y[i + 1]:
            k = (y[i + 1] - y[i]) / (x[i + 1] - x[i])
            x0 = x[i] + (y0 - y[i]) / k
            return x0
    else:
        raise ValueError('【Error】未找到交点-1')

def get_y(x: list, y: list, x0: float, error: bool=True) -> float:
    """获得竖线x=x0与给定曲线的交点纵坐标

    Args:
        x (list): 输入曲线的横坐标序列
        y (list): 输入曲线的纵坐标序列
        x0 (float): 竖直线x = x0
        error (boo, optional): 若x0超出范围，抛出异常or返回None

    Returns:
        float: 曲线与竖线交点纵坐标
    """
    # 获得x=x0与曲线的交点
    if x0 < min(x):
        if error:
            raise ValueError(f'【Error】x0 < min(x) ({x0} < {min(x)})')
        else:
            return None
    if x0 > max(x):
        if error:
            raise ValueError(f'【Error】x0 > max(x) ({x0} > {max(x)})')
        else:
            return None
    for i in range(len(x) - 1):
        if x[i] == x0:
            y0 = y[i]
            return y0
        elif x[i] < x0 <= x[i + 1]:
            k = (y[i + 1] - y[i]) / (x[i + 1] - x[i])
            y0 = k * (x0 - x[i]) + y[i]
            return y0
    else:
        raise ValueError('【Error】未找到交点-2')

    
def get_percentile_line(
        all_x: list[list],
        all_y: list[list],
        p: float,
        n: int=None,
        x: list | np.ndarray=None
    ) -> tuple[np.ndarray, np.ndarray]:
    """计算IDA曲线簇的百分位线

    Args:
        all_x (list[list]): 所有独立IDA的横坐标
        all_y (list[list]): 所有独立IDA的纵坐标
        p (float): 百分位值
        n (int): 输出的百分位线横坐标的点数量
        x (list | np.ndarray, optional): 百分位线横坐标，默认为None，即取最大适用范围

    Returns:
        tuple[np.ndarray, np.ndarray]: 百分位线的横坐标、纵坐标
    """
    # 计算百分位线
    if x is None:
        x1 = min([min(i) for i in all_x])
        x2 = max([max(i) for i in all_x])
        x = np.linspace(x1, x2, n)  # 百分位线横坐标
    y = []  # 百分位线纵坐标
    for _, xi in enumerate(x):
        # xi: int, yi: list
        yi = []
        for _, (line_x, line_y) in enumerate(zip(all_x, all_y)):
            res = get_y(line_x, line_y, xi, False)
            if res is not None:
                yi.append(res)
        y_percentile = np.percentile(yi, p)
        y.append(y_percentile)
    y = np.array(y)
    return x, y

def _get_mean_std_line(all_x: list[list], all_y: list[list], n: int) -> tuple[np.ndarray, np.ndarray]:
    """计算IDA曲线簇的标准差线

    Args:
        all_x (list[list]): 所有独立IDA的横坐标
        all_y (list[list]): 所有独立IDA的纵坐标
        n (int): 输出的STD线横坐标的点数量

    Returns:
        tuple[np.ndarray, np.ndarray]: STD线的横坐标、纵坐标
    """
    # 计算百分位线
    x1 = min([min(i) for i in all_x])
    x2 = max([max(i) for i in all_x])
    x = np.linspace(x1, x2, n)  # 百分位线横坐标
    y_mean = []  # 均值线纵坐标
    y_std = []  # 标准差线纵坐标
    for _, xi in enumerate(x):
        # xi: int, yi: list
        yi = []
        for _, (line_x, line_y) in enumerate(zip(all_x, all_y)):
            res = get_y(line_x, line_y, xi, False)
            if res is not None:
                yi.append(res)
        y_mean_ = np.mean(yi)
        y_std_ = np.std(yi, ddof=0)
        y_mean.append(y_mean_)
        y_std.append(y_std_)
    y_mean = np.array(y_mean)
    y_std = np.array(y_std)
    return x, y_mean, y_std


class FragilityAnalysis():
    available_EDP_types = ['IM', 'IDR', 'DCF', 'PFV', 'PFA', 'RIDR', 'RoofIDR' ,'Shear', 'beamHinge', 'colHinge', 'panelZone']  # 允许的DM类型

    def __init__(self, root: str | Path, EDP_types: list[str],
                 collapse_limit: float=0.1, additional_items: list[str]=None):
        """地震易损性、倒塌易损性分析

        Args:
            root (str | Path): 待读取数据的文件夹的路径
            EDP_types (list[str]): 工程需求参数类型
            * [IDR] - 层间位移角
            * [DCF] - 层间变形集中系数
            * [PFV] - 层间速度
            * [PFA] - 楼层绝对加速度
            * [RIDR] - 残余层间位移角
            * [RoofIDR] - 屋顶层间位移角
            * [Shear] - 层间剪力
            * [beamHinge] - 最大梁铰变形
            * [colHinge] - 最大柱铰变形
            * [panelZone] - 最大节点域变形
            * additional_item - 其他项\n
            collapse_limit (float, optional): 倒塌极限层间位移角，如果给定DM对应的层间位移角角大于给定值，
            则不统计，可设为0.1或0.15，默认为0.1\n
            additional_items (list[str], optional): 其他的要读取的项
        """
        for EDP_type in EDP_types:
            if not EDP_type in self.available_EDP_types:
                raise ValueError(f'`{EDP_type}`不是可用的类型')
        self.Calc_collapse = False  # 是否有进行倒塌易损性计算
        self.Calc_p = False  # 是否有进行EDP超越概率计算
        self.root = Path(root)
        self.EDP_types = EDP_types
        self.collapse_limit = collapse_limit
        self.additional_items = additional_items
        if self.additional_items:
            self.EDP_types += self.additional_items
        self.has_risk_data = False  # 是否进行了风险评估
        self._init_variables()
        self._check_file()
        self._read_file()
        # ResultFolder = model + '_new'
        # self.init_ReadFile(ResultFolder, EDP_type)

    def _init_variables(self):
        """初始化参数"""
        # calc_IDA
        self.IM_scatter: dict[str, list] = {}  # IDA曲线散点(包含临界倒塌点)
        self.DM_scatter: dict[str, list] = {}
        self.IM_scatter1: dict[str, list] = {}  # IDA曲线散点(不包含临界倒塌点)
        self.DM_scatter1: dict[str, list] = {}
        self.IM_lines: dict[str, list[list]] = {}  # IDA曲线
        self.DM_lines: dict[str, list[list]] = {}
        self.ln_IM:dict[str, list] = {}
        self.ln_DM:dict[str, list] = {}
        self.pct_x, self.pct_16, self.pct_50, self.pct_84, self.pct_100, self.mean, self.std = {}, {}, {}, {}, {}, {}, {}
        # frag_curve
        self.DM_limits: dict[str, float] = {}  # 各DM的最大值，超过该值则不统计
        self.AB: dict[str, tuple[float, float]] = {}  # ln(DM) = A + B * ln(IM)
        self.R2: dict[str, float] = {}  # 拟合优度
        self.betaD: dict[str, float] = {}  # 地震需求不确定性(根据PSDM拟合结果计算)
        self.ln_IM_line, self.ln_DM_line = {}, {}  # ln(IM)和ln(DM)的拟合曲线
        self.x_frag: dict[str, np.ndarray] = {}  # 易损性曲线横坐标
        self.y_frag: dict[str, list[np.ndarray]] = {}  # 易损性曲线纵坐标(多个)
        self.DS: dict[str, dict[str, float]] = {}  # 损伤状态及对应标签
        self.beta: dict[str, float] = {}  # 不确定性系数
        self.info: dict[str, str] = {}  # 记录拟合参数
        # exceedance_probability
        self.exceed_mean: dict[str, dict[str, float]]  = {}  # 超越概率均值
        self.exceed_std: dict[str, dict[str, float]]  = {}  # 超越概率标准差
        self.exceed_pct50: dict[str, dict[str, float]] = {}  # 中位值
        self.exceed_x: dict[str, dict[str, list[float]]] = {}  # 超越概率点的横坐标
        self.exceed_y: dict[str, dict[str, np.ndarray]] = {}  # 超越概率点的纵坐标
        self.exceed_x_fit: dict[str, dict[str, np.ndarray]] = {}  # 超越概率点的拟合曲线横坐标
        self.exceed_y_fit: dict[str, dict[str, np.ndarray]] = {}  # 超越概率点的拟合曲线纵坐标
        self.exceed_y_fixedBeta: dict[str, dict[str, np.ndarray]] = {}  # 采用固定不确定性导致的标准差的易损性曲线
        self.DM_values: dict[str, dict[str, float]] = {}  # 超越概率计算所需的DM值
        self.DM_has_fixed_beta: dict[str, dict[str, float]] = {}  # 指定了固定不确定性beta_TOT的DM类型
        # manual_probability
        self.hazard_curve: np.ndarray = None  # 地震危险性曲线
        self.risk_figs: dict[str, Figure] = {}
        self.risk_text: dict[str, str] = {}
        self.risk_EDP_hazard_curves: dict[str, np.ndarray] = {}
        # collapse_evaluation
        self.has_collapse_data = False  # 是否有倒塌数据
        self.df_clps_frag = pd.DataFrame(None)  # 倒塌易损性曲线
        self.collapse_intensity = {}  # 倒塌强度统计指标
        self.all_CMR: dict[str, float] = {}  # CMR
        self.collapse_risk = {}  # 倒塌风险
        # 可视化
        self.all_figures_1: dict[str, tuple[Figure, list[list[Axes]]]] = {} # 所有易损性相关图
        self.all_figures_2: dict[str, tuple[Figure, list[list[Axes]]]] = {}  # 所有概率风险相关图
        for EDP_type in self.EDP_types[::-1]:
            fig, axes = plt.subplots(nrows=1, ncols=3, figsize=(16, 5))
            fig.suptitle(f'Risk Analysis of `{EDP_type}`')
            self.all_figures_2[EDP_type] = (fig, axes)
            plt.tight_layout(w_pad=3)
        for EDP_type in self.EDP_types[::-1]:
            fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(16, 12))
            fig.suptitle(f'Fragility Analysis of `{EDP_type}`')
            self.all_figures_1[EDP_type] = (fig, axes)
            plt.tight_layout(pad=3)

    def _check_file(self):
        # 数据初步检查
        list_ = ['ground_motions', 'Nstory', 'Nbay', 'running_case']
        for file in list_:
            if not Path.exists(self.root/f'{file}.dat'):
                raise ValueError('【Error】未找到{}'.format(self.root/f'{file}.dat'))
        self.GM_names = np.loadtxt(self.root/'ground_motions.dat', dtype=str, ndmin=1).tolist()  # 地震动名
        self.GM_N = len(self.GM_names)
        self.Nstory = int(np.loadtxt(self.root/'Nstory.dat'))  # 楼层数
        self.Nbay = int(np.loadtxt(self.root/'Nbay.dat'))  # 跨数
        with open(self.root/'notes.dat', 'r') as f:
            self.notes = f.read()
        self.running_case = str(np.loadtxt(self.root / 'running_case.dat', dtype=str))
        if self.running_case != 'IDA':
            raise ValueError('【Error】本程序仅支持处理IDA结果数据')
        logger.success('通过数据文件检查')

    def _read_file(self):
        """读取结果文件"""
        # 读取数据
        self.data: list[pd.DataFrame] = []  # 包含所有时程结果的最值
        columns = self.available_EDP_types
        if self.additional_items:
            columns += self.additional_items
        for idx_gm in range(self.GM_N):
            # 遍历地震动
            gm_name = self.GM_names[idx_gm]
            print(f'  正在读取: {gm_name}    \r', end='')
            num =  1
            df = pd.DataFrame(columns=columns)
            while True:
                # 遍历每个动力增量
                folder = self.root / f'{gm_name}_{num}'
                if not Path.exists(folder):
                    break
                line = []
                IM = np.loadtxt(folder/'Sa.out')  # 地震动强度指标
                IM = round(float(IM), 6)
                line.append(IM)
                if (folder/'层间位移角.out').exists():
                    IDR = np.loadtxt(folder/'层间位移角.out')  # 层间位移角
                    line.append(np.max(np.abs(IDR)))
                else:
                    line.append(0)
                if (folder/'DCF.out').exists():
                    DCF = np.loadtxt(folder/'DCF.out')  # 层间变形集中系数
                    line.append(np.max(np.abs(DCF)))
                else:
                    line.append(0)
                if (folder/'层速度.out').exists():
                    PFV = np.loadtxt(folder/'层速度.out')  # 层间相对速度
                    line.append(np.max(np.abs(PFV)))
                else:
                    line.append(0)
                if (folder/'层加速度(g).out').exists():
                    PFA = np.loadtxt(folder/'层加速度(g).out')  # 层间相对加速度
                    line.append(np.max(np.abs(PFA)))
                else:
                    line.append(0)
                if (folder/'残余层间位移角.out').exists():
                    RIDR = np.loadtxt(folder/'残余层间位移角.out')  # 残余层间位移角
                    line.append(np.max(np.abs(RIDR)))
                else:
                    line.append(0)
                if (folder/'屋顶层间位移角.out').exists():
                    RoofIDR = np.loadtxt(folder/'屋顶层间位移角.out')  # 屋顶层间位移角
                    line.append(np.max(np.abs(RoofIDR)))
                else:
                    line.append(0)
                if (folder/'楼层剪力(kN).out').exists():
                    Shear = np.loadtxt(folder/'楼层剪力(kN).out')  # 层间剪力
                    line.append(np.max(np.abs(Shear)))
                else:
                    line.append(0)
                if (folder/'梁铰变形.out').exists():
                    beamHinge = np.loadtxt(folder/'梁铰变形.out')  # 梁铰变形
                    line.append(np.max(np.abs(beamHinge)))
                else:
                    line.append(0)
                if (folder/'柱铰变形.out').exists():
                    colHinge = np.loadtxt(folder/'柱铰变形.out')  # 柱铰变形
                    line.append(np.max(np.abs(colHinge)))
                else:
                    line.append(0)
                if (folder/'节点域变形.out').exists():
                    panelZone = np.loadtxt(folder/'节点域变形.out')  # 节点域变形
                    line.append(np.max(np.abs(panelZone)))
                else:
                    line.append(0)
                if self.additional_items:
                    for item in self.additional_items:
                        if not (folder/f'{item}.out').exists():
                            raise FileExistsError(f'无法找到文件:', str((folder/f'{item}.out').absolute()))
                        data = np.loadtxt(folder/f'{item}.out')  # 其他的要读取的项
                        line.append(np.max(np.abs(data)))
                df.loc[len(df)] = line
                num += 1
            df = df.sort_values(by='IM')  # 每条地震波按IM列的大小排序
            df.index = [i for i in range(1, len(df) + 1)]  # 重新指定index
            for idx, line in df.iterrows():
                if line['IDR'] >= self.collapse_limit:
                    break
            df = df[df.index<=idx]  # 保留临界倒塌点，删除之后的数据
            self.data.append(df)
        logger.success('已读取数据')

    def calc_IDA(self,
            EDP_type: str,
            DM_limit: float=None,
            DM_display_limit: float=None,
            density: int=300,
            marked_idx: int=None
        ):
        """计算IDA曲线

        Args:
            EDP_type (str): 工程需求参数类型
            DM_limit (float, optional): DM的上限值，超过该值则不统计，默认无上限
            DM_display_limit (float, optional): 绘图时DM的上限值，默认为无穷大，该值只对绘图有影响，不会影响计算结果
            density (int, optional): IDA曲线蔟分位线的点数量，默认为300
            marked_idx (int, optional): 标记的IDA曲线的索引用于画图时高亮显示，默认为None
        """
        print(f'  正在计算`{EDP_type}`类型的IDA曲线      \r', end='')
        IM_scatter, DM_scatter = [], []  # 所有IDA曲线的散点坐标
        IM_scatter1, DM_scatter1 = [], []  # 所有IDA曲线的散点坐标（不包含倒塌点）
        IM_lines, DM_lines = [], []  # IDA曲线簇(多条)
        if DM_limit is not None:
            self.DM_limits[EDP_type] = DM_limit
        for df in self.data:
            IM_i = df['IM'].to_numpy()
            DM_i = df[EDP_type].to_numpy()
            if (DM_limit is not None) and (np.max(DM_i) > DM_limit):
                # 剔除超过DM_limit的点
                IM_i = np.append(IM_i[DM_i<=DM_limit], IM_i[DM_i>DM_limit][0])
                DM_i = np.append(DM_i[DM_i<=DM_limit], DM_i[DM_i>DM_limit][0])
            IM_i, DM_i = list(IM_i), list(DM_i)
            IM_scatter += IM_i  # 包含临界倒塌点
            DM_scatter += DM_i
            IM_scatter1 += IM_i[:-1]  # 不包含临界倒塌点
            DM_scatter1 += DM_i[:-1]
            IM_lines.append([0] + IM_i)
            DM_lines.append([0] + DM_i)
        pct_x, pct_16 = get_percentile_line(DM_lines, IM_lines, p=16, n=density)
        pct_x, pct_50 = get_percentile_line(DM_lines, IM_lines, p=50, n=density)
        pct_x, pct_84 = get_percentile_line(DM_lines, IM_lines, p=84, n=density)
        pct_x, pct_100 = get_percentile_line(DM_lines, IM_lines, p=100, n=density)
        _, mean, std = _get_mean_std_line(DM_lines, IM_lines, n=density)
        self.IM_scatter[EDP_type] = IM_scatter
        self.DM_scatter[EDP_type] = DM_scatter
        self.IM_scatter1[EDP_type] = IM_scatter1
        self.DM_scatter1[EDP_type] = DM_scatter1
        self.IM_lines[EDP_type] = IM_lines
        self.DM_lines[EDP_type] = DM_lines
        self.pct_x[EDP_type] = pct_x
        self.pct_16[EDP_type] = pct_16
        self.pct_50[EDP_type] = pct_50
        self.pct_84[EDP_type] = pct_84
        self.pct_100[EDP_type] = pct_100
        self.mean[EDP_type] = mean
        self.std[EDP_type] = std
        # 画图 IDA曲线
        fig, axes = self.all_figures_1[EDP_type]
        ax = axes[0][0]
        label1 = 'Uncollapsed points'
        label2 = 'Collapsed points'
        for i, (x, y) in enumerate(zip(self.DM_lines[EDP_type], self.IM_lines[EDP_type])):
            if marked_idx == i:
                ax.plot(x, y, color='blue', markersize=6, label=f'idx={i}', zorder=9999)
            else:
                ax.plot(x, y, color='#BFBFBF', markersize=4)
            ax.scatter(x[:-1], y[:-1], color='#E59EDD', zorder=8999, s=7, label=label1)  # 未倒塌点
            ax.scatter(x[-1], y[-1], color='red', zorder=8999, s=10, label=label2)  # 临界倒塌点
            label1 = None
            label2 = None
        ax.plot(self.pct_x[EDP_type], self.pct_16[EDP_type], label='16%', color='green', linewidth=3, linestyle='--')
        ax.plot(self.pct_x[EDP_type], self.pct_50[EDP_type], label='50%', color='green', linewidth=3)
        ax.plot(self.pct_x[EDP_type], self.pct_84[EDP_type], label='84%', color='green', linewidth=3, linestyle='--')
        ax.set_title(f'IDA curves ({EDP_type})')
        ax.legend()
        if EDP_type == 'IDR':
            ax.set_xlim(0, self.collapse_limit)
        else:
            ax.set_xlim(0)
        if EDP_type in self.DM_limits.keys():
            ax.set_xlim(right=self.DM_limits[EDP_type])
        ax.set_ylim(0)
        if DM_display_limit is not None:
            ax.set_xlim(right=DM_display_limit)
        ax.set_xlabel('DM')
        ax.set_ylabel('IM')
        logger.success(f'已计算`{EDP_type}`类型的IDA曲线')

    def frag_curve(
            self, EDP_type: str,
            DS: dict[str, float],
            beta: list[float | Literal['betaD']] | float,
            IM_limit: float=None,
            density: float=1000,
            DM_min: float=None,
            DM_max: float=None,
        ):
        """对概率需求模型进行拟合，计算易损性曲线。  

        Args:
            EDP_type (str): 工程需求参数名称
            DS (dict[str, float]): 损伤状态标签及对应数值
            beta (list[float | Literal['betaD']] | float): 不确定性参数
            IM_limit (float, optional): 易损性曲线的最大横坐标范围，默认为None，取倒塌时84%分位地震动强度的1.8倍
            density (float, optional): 易损性曲线的点密度，默认为1000
            DM_min (float, optional): DM下限值，默认为None表示不限制
            DM_max (float, optional): DM上限值，默认为None表示不限制

        注1: 
        -----
        通过该方法计算的易损性函数具有分解表达式:  
        FR(x) = P[D≥C|IM=x] = DCF((ln(mD(x) - ln(mC)) / sqrt(beta))),  
        其中, `mD(x)`是中值, `mC`是能力值
        根据概率地震需求模型, IM和DM服从对数线性关, 即  
        mD(x) = ax^b  

        注2:
        ------
        不确定性参数`beta`可填多个值，也可以填`'betaD'`表示使用PSDM的对数标准差，作为需求不确定性系数
        多个`beta`值将按下式进行叠加: `beta = sqrt(beta_1^2 + beta_2^2 + beta_3^2 + ...)`
        例如：  
        >>> beta = [0.2, 0.3, 'betaD']
        表示beta = sqrt(0.2^2 + 0.3^2 + betaD^2)，其中betaD根据PSDM的拟合结果计算。
        也可只传入单个不确定值，例如：  
        >>> beta = 0.4  
        表示beta = 0.4。  

        通常, `beta`可按下式计算：  
        beta = sqrt(betaD^2 + betaC^2)  
        其中`betaD`和`betaC`是需求和能力的不确定性系数, `betaD`按下式计算:  
        betaD = sqrt(sum_{i=0}^N (lnDi - lnmD)^2 / (N - 2))  
        """
        if not EDP_type in self.EDP_types:
            raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
        self.DS[EDP_type] = DS
        # EDP_type: 损伤指标名称, DS: 损伤状态值
        print(f'  正在计算`{EDP_type}`类型的易损性曲线和概率需求模型      \r', end='')
        IM = np.array(self.IM_scatter1[EDP_type])
        DM = np.array(self.DM_scatter1[EDP_type])
        if DM_min is not None and DM_max is not None and DM_min > DM_max:
            raise ValueError(f'DM下限({DM_min})不可大于上限({DM_max})')
        if DM_min is not None or DM_max is not None:
            mask = np.ones_like(DM, dtype=bool)
            if DM_min is not None:
                mask &= DM >= DM_min
            if DM_max is not None:
                mask &= DM <= DM_max
            removed_points = int(len(DM) - np.sum(mask))
            IM = IM[mask]
            DM = DM[mask]
            if len(DM) < 3:
                raise ValueError('过滤后的DM散点数量不足（至少需要3个点）')
            if removed_points:
                logger.warning(
                    f'`{EDP_type}`类型因DM范围过滤掉{removed_points}个散点 '
                    f'(DM_min={DM_min}, DM_max={DM_max})'
                )
        # 地震易损性曲线的横坐标最大范围(取为84分位线的最大IM值的3倍)
        if IM_limit is None:
            IM2_for_frag_curve = get_y(self.pct_x[EDP_type], self.pct_100[EDP_type], max(DS.values())) * 1.5
        else:
            IM2_for_frag_curve = IM_limit
        # 概率需求模型曲线的横坐标最大范围(取为最大IM值)
        ln_IM, ln_DM = np.log(IM), np.log(DM)
        ln_IM_line = np.linspace(min(ln_IM), max(ln_IM), 1001)
        mean_ln_IM = np.mean(ln_IM)
        mean_ln_DM = np.mean(ln_DM)
        B = np.sum((ln_IM - mean_ln_IM) * (ln_DM - mean_ln_DM)) / np.sum((ln_IM - mean_ln_IM) ** 2)
        A = mean_ln_DM - B * mean_ln_IM
        ln_DM_pre = A + B * ln_IM
        SSR = np.sum((ln_DM - ln_DM_pre) ** 2)
        SST = np.sum((ln_DM - np.mean(ln_DM_pre)) ** 2)
        R2 = 1 - (SSR / SST)
        N = len(DM)
        betaD = np.sqrt(np.sum((ln_DM - ln_DM_pre) ** 2) / (N - 2))  # 对数标准差
        beta_total = 0  # 总不确定性
        try:
            iter(beta)
        except TypeError:
            beta = [beta]
        for beta_i in beta:
            if beta_i == 'betaD':
                beta_total += betaD ** 2
            else:
                beta_total += beta_i ** 2
        beta_total = float(np.sqrt(beta_total))
        x_frag = np.linspace(0.001, IM2_for_frag_curve, density)  # 易损性曲线x轴
        y_frags: list[np.ndarray] = []  # 易损性曲线y轴
        for _, DS_i in DS.items():
            y_frags.append(norm.cdf((A + B * np.log(x_frag) - np.log(DS_i)) / beta_total, 0, 1))  # 易损性曲线y轴
        # DM和IM的对数关系
        ln_DM_line = A + B * ln_IM_line
        self.AB[EDP_type] = (A, B)
        self.R2[EDP_type] = R2
        self.beta[EDP_type] = beta_total  # 总不确定性
        self.betaD[EDP_type] = betaD  # 地震需求不确定性
        self.ln_IM_line[EDP_type] = ln_IM_line
        self.ln_DM_line[EDP_type] = ln_DM_line
        self.x_frag[EDP_type] = x_frag
        self.y_frag[EDP_type] = y_frags
        self.ln_IM[EDP_type] = ln_IM
        self.ln_DM[EDP_type] = ln_DM
        text = f'类型`{EDP_type}`\n\n概率模型需求参数(ln(DM) = A + B * ln(IM))：\n'
        text += f'A = {A:.6f}\n'
        text += f'B = {B:.6f}\n'
        text += f'R2 = {R2:.6f}\n'
        text += f'beta_D = {betaD:.6f}\n'
        text += f'beta_total = {beta_total:.6f}\n\n'
        self.info[EDP_type] = text
        # 画图: 概率需求曲线
        fig, axes = self.all_figures_1[EDP_type]
        ax = axes.flatten()[0]
        x = np.array(self.pct_x[EDP_type])
        y = (x / np.exp(A)) ** (1 / B)
        ax.plot(x, y, color='red', label='PSDM')
        ax.legend()
        # 画图: PSDM拟合
        ax = axes.flatten()[1]
        ax.plot(ln_IM, ln_DM, 'o')  # 散点
        ax.plot(ln_IM_line, ln_DM_line, 'red', label=f'ln(DM) = {A:.4f} + {B:.4f} * ln(IM)\n$R^2$ = {R2:.4f}')
        ax.legend()
        ax.set_title('ln(DM) - ln(IM)')
        ax.set_xlabel('ln(IM)')
        ax.set_ylabel('ln(DM)')
        # 画图: 易损性曲线
        ax = axes.flatten()[2]
        for i, (label, DS_i) in enumerate(DS.items()):
            ax.plot(x_frag, y_frags[i], label=f'{label} ({DS_i})')
        ax.legend()
        ax.set_title('Fragility curves')
        ax.set_xlim(0)
        ax.set_ylim(0, 1)
        ax.set_xlabel('IM')
        ax.set_ylabel('Exceeding probability')
        logger.success('已完成易损性函数计算和概率需求模型的拟合')

    def exceedance_probability(self,
            EDP_type: str,
            DM_value: float | dict[str, float],
            beta: list[float] | float=None,
            *args, **kwargs
        ):
        """计算所有损伤指标超越一个或多个阈值的概率

        注:
        -----
        该方法直接以`DM_value`所在的水平线切割所有IDA曲线，获得对应的IM散点，
        并构建整体的倒塌易损性函数: FR(x) = P[D|IM=x] = Φ((ln(mD(x)) - ln(mC)) / β)。
        其中`mD(x)`为中值需求，`mC`为能力，`β`为需求总不确定性。

        β说明:
        ------
        `beta`可传入单个值或多个值(列表)，程序会以`β = sqrt(β1^2 + β2^2 + …)`进行组合作为体系总不确定性。
        无论是否传入`beta`，程序都会自动计算一次`beta = std(IM(D), ddof=1)`，即指定需求下IM的标准差。
        如果提供了`beta`，则同时输出“指定β”和“统计β”下的曲线。

        注意:
        -----
        当`EDP_type`为`IDR`时，基于该方法得到的倒塌易损性曲线将用于`collapse_evaluation`。
        多个`DM_value`会依次计算，并在同一张图上展示。

        Args:
            EDP_type (str): 工程需求参数类型。
            DM_value (float | dict[str, float]): 单个阈值，或形如`{'DS-1': 0.01, ...}`的多阈值。
            beta (list[float] | float, optional): 指定的体系总不确定性，默认None，仅考虑记录离散性。
        """
        internal_call: bool = kwargs.get('internal_call', False)
        if EDP_type not in self.EDP_types:
            raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
        if isinstance(DM_value, dict):
            DS_items = list(DM_value.items())
        else:
            DS_items = [(self._format_exceedance_label(DM_value), DM_value)]
        if internal_call:
            if len(DS_items) != 1:
                raise ValueError('内部调用仅允许单个阈值的超越概率计算')
            label, target_DM = DS_items[0]
            stats = self._compute_exceedance_stats(EDP_type, target_DM, beta, label=label, suppress_log=True)
            if stats is None:
                return None
            df = self._build_exceedance_dataframe(
                stats['scatter_x'], stats['scatter_y'],
                stats['fit_x'], stats['fit_y'], stats['fixed_y']
            )
            return stats['theta'], stats['beta_calc'], stats['fit_x'], stats['fit_y'], stats['fixed_y'], stats['scatter_x'], df
        for label, target_DM in DS_items:
            stats = self._compute_exceedance_stats(EDP_type, target_DM, beta, label=label)
            if stats is None:
                continue
            text_summary = self._format_exceedance_summary(EDP_type, label, target_DM, stats)
            print(text_summary)
            self._store_exceedance_results(EDP_type, label, target_DM, stats, text_summary)
        self._refresh_exceedance_plot(EDP_type)
        logger.success('已计算超越概率曲线')

    def _compute_exceedance_stats(self,
            EDP_type: str,
            DM_value: float,
            beta: list[float] | float | None,
            label: str | None=None,
            suppress_log: bool=False,
        ) -> dict | None:
        IM_lines, DM_lines = self.IM_lines[EDP_type], self.DM_lines[EDP_type]
        if DM_value < min([min(i) for i in DM_lines]):
            raise ValueError(f'`{EDP_type}`类型的`{DM_value}`小于所有IDA曲线的最小DM值({min([max(i) for i in DM_lines])})')
        if DM_value > max([max(i) for i in DM_lines]):
            raise ValueError(f'`{EDP_type}`类型的`{DM_value}`大于所有IDA曲线的最大DM值({max([max(i) for i in DM_lines])})')
        IM_points = []
        for i in range(self.GM_N):
            IM_line, DM_line = IM_lines[i], DM_lines[i]
            try:
                y = get_y(DM_line, IM_line, DM_value)
                IM_points.append(y)
            except ValueError:
                logger.warning(f'`{EDP_type}`类型的`{DM_value}`不在第{i+1}条IDA曲线的DM范围({min(DM_line)}, {max(DM_line)})')
        if len(IM_points) < self.GM_N:
            logger.error(f'`{EDP_type}`类型的超越概率统计失败')
            return None
        exceed_mean = float(np.mean(IM_points))
        exceed_std = float(np.std(IM_points, ddof=1))
        exceed_pct50 = float(np.percentile(IM_points, 50))
        exceed_x: list[float] = sorted(IM_points)
        exceed_y = np.array([i/self.GM_N for i in range(1, self.GM_N+1)])
        if beta is not None:
            try:
                iter(beta)
            except TypeError:
                beta = [beta]
            beta_total = 0.0
            for beta_i in beta:
                beta_total += beta_i ** 2
            beta_total = float(np.sqrt(beta_total))
        else:
            beta_total = None
        epsilon = 1e-4
        z = norm.ppf(np.clip(exceed_y, epsilon, 1 - epsilon))
        lnIM = np.log(exceed_x)
        A = np.vstack([z, np.ones_like(z)]).T
        beta_calc, ln_theta = np.linalg.lstsq(A, lnIM, rcond=None)[0]
        theta = np.exp(ln_theta)
        IM1, IM2 = 0.001, max(exceed_x) * 1.2
        exceed_x_fit = np.linspace(IM1, IM2, 1001)
        exceed_y_fit = norm.cdf(np.log(exceed_x_fit / theta) / beta_calc, 0, 1)
        if beta_total is not None and not suppress_log:
            if label is not None:
                logger.success(f'为`{EDP_type}`类型的`{label}`指定了体系总不确定性 beta_total = {beta_total}')
            else:
                logger.success(f'为`{EDP_type}`类型指定了体系总不确定性 beta_total = {beta_total}')
        beta_fixed = beta_total if beta_total is not None else beta_calc
        exceed_y_fixedBeta = norm.cdf(np.log(exceed_x_fit / theta) / beta_fixed, 0, 1)
        return {
            'mean': exceed_mean,
            'std': exceed_std,
            'pct50': exceed_pct50,
            'scatter_x': exceed_x,
            'scatter_y': exceed_y,
            'fit_x': exceed_x_fit,
            'fit_y': exceed_y_fit,
            'fixed_y': exceed_y_fixedBeta,
            'theta': theta,
            'beta_calc': beta_calc,
            'beta_fixed': beta_fixed,
            'has_fixed_beta': beta_total is not None,
        }

    @staticmethod
    def _build_exceedance_dataframe(
            exceed_x: list[float],
            exceed_y: np.ndarray,
            exceed_x_fit: np.ndarray,
            exceed_y_fit: np.ndarray,
            exceed_y_fixedBeta: np.ndarray,
        ) -> pd.DataFrame:
        N_lines = max(len(exceed_x), len(exceed_x_fit), len(exceed_x_fit))
        df = pd.DataFrame(None)
        df['Sa_scatters'] = pd.Series(exceed_x).reindex(range(N_lines))
        df['IM_scatters'] = pd.Series(exceed_y).reindex(range(N_lines))
        df['Sa_fixedBeta'] = pd.Series(exceed_x_fit).reindex(range(N_lines))
        df['IM_fixedBeta'] = pd.Series(exceed_y_fixedBeta).reindex(range(N_lines))
        df['Sa_fit'] = pd.Series(exceed_x_fit).reindex(range(N_lines))
        df['IM_fit'] = pd.Series(exceed_y_fit).reindex(range(N_lines))
        return df

    def _store_exceedance_results(self,
            EDP_type: str,
            label: str,
            DM_value: float,
            stats: dict,
            text_summary: str,
        ) -> None:
        self.exceed_mean.setdefault(EDP_type, {})[label] = stats['mean']
        self.exceed_std.setdefault(EDP_type, {})[label] = stats['std']
        self.exceed_pct50.setdefault(EDP_type, {})[label] = stats['pct50']
        self.exceed_x.setdefault(EDP_type, {})[label] = stats['scatter_x']
        self.exceed_y.setdefault(EDP_type, {})[label] = stats['scatter_y']
        self.exceed_x_fit.setdefault(EDP_type, {})[label] = stats['fit_x']
        self.exceed_y_fit.setdefault(EDP_type, {})[label] = stats['fit_y']
        self.exceed_y_fixedBeta.setdefault(EDP_type, {})[label] = stats['fixed_y']
        self.DM_values.setdefault(EDP_type, {})[label] = DM_value
        if stats['has_fixed_beta']:
            self.DM_has_fixed_beta.setdefault(EDP_type, {})[label] = stats['beta_fixed']
        elif EDP_type in self.DM_has_fixed_beta and label in self.DM_has_fixed_beta[EDP_type]:
            self.DM_has_fixed_beta[EDP_type].pop(label)
            if not self.DM_has_fixed_beta[EDP_type]:
                self.DM_has_fixed_beta.pop(EDP_type)
        self.info.setdefault(EDP_type, '')
        self.info[EDP_type] += text_summary + '\n'


    def _refresh_exceedance_plot(self, EDP_type: str) -> None:
        if EDP_type not in self.all_figures_1 or EDP_type not in self.exceed_x:
            return
        fig, axes = self.all_figures_1[EDP_type]
        ax = axes.flatten()[3]
        ax.cla()
        ax.set_title(f'Exceedance probability ({EDP_type})')
        ax.set_ylabel('Exceedance probability')
        ax.set_xlabel('IM')
        for label, DM_value in self.DM_values.get(EDP_type, {}).items():
            x_fit = self.exceed_x_fit[EDP_type][label]
            y_fit = self.exceed_y_fit[EDP_type][label]
            ax.plot(x_fit, y_fit, label=f'{label} ({DM_value}) fit')
            ax.plot(self.exceed_x[EDP_type][label], self.exceed_y[EDP_type][label], 'o', label=f'{label} scatter')
            if label in self.DM_has_fixed_beta.get(EDP_type, {}):
                ax.plot(x_fit, self.exceed_y_fixedBeta[EDP_type][label], linestyle='--', label=f'{label} fixed beta')
        ax.set_xlim(0)
        ax.set_ylim(0)
        if ax.has_data():
            ax.legend()

    @staticmethod
    def _format_exceedance_summary(
            EDP_type: str,
            label: str,
            DM_value: float,
            stats: dict,
        ) -> str:
        text = f'`{EDP_type}`超越{DM_value}({label})的概率特征：\n'
        text += f'均值：{stats["mean"]:.6f}\n'
        text += f'标准差：{stats["std"]:.6f}\n'
        text += f'中位值：{stats["pct50"]:.6f}'
        return text


    @staticmethod
    def _format_exceedance_label(DM_value: float) -> str:
        return f'DM>{DM_value:g}'

    @staticmethod
    def _parse_hazard_period(path: Path) -> float:
        try:
            return float(Path(path).stem)
        except ValueError as exc:
            raise ValueError(f"无法从文件名 `{Path(path).name}` 解析周期") from exc

    @staticmethod
    def _read_hazard_curve(path: Path) -> tuple[np.ndarray, np.ndarray]:
        data = np.loadtxt(path)
        if data.ndim != 2 or data.shape[1] < 2:
            raise ValueError(f"{path} 不是两列格式 (Sa, λ(IM))")
        if np.any(data[:, 0] <= 0) or np.any(data[:, 1] <= 0):
            raise ValueError(f"{path} 包含非正值，无法用于 log-log 插值")
        order = np.argsort(data[:, 0])
        return data[order, 0], data[order, 1]

    def interpolate_hazard_curve(
        self,
        hazard_dir: str | Path,
        target_period: Real | str | Path,
        resample_points: int = 40,
    ) -> np.ndarray:
        """根据结构基本周期自动插值得到目标危险性曲线（参考 data/HC.py）。"""
        if resample_points < 2:
            raise ValueError("resample_points 至少为 2")
        hazard_dir = Path(hazard_dir)
        if not hazard_dir.exists():
            raise FileNotFoundError(f"未找到灾害曲线文件夹：{hazard_dir}")

        files = [
            f
            for pattern in ("*.txt", "*.out")
            for f in hazard_dir.glob(pattern)
        ]
        if not files:
            raise FileNotFoundError(f"文件夹中未找到任何灾害曲线：{hazard_dir}")
        files = sorted(files, key=self._parse_hazard_period)
        periods = np.array([self._parse_hazard_period(f) for f in files], dtype=float)

        if isinstance(target_period, Real):
            T_target = float(target_period)
        else:
            period_path = Path(target_period)
            if not period_path.exists():
                raise FileNotFoundError(f"无法找到周期文件：{period_path}")
            period_vals = np.loadtxt(period_path, ndmin=1, dtype=float)
            T_target = float(np.atleast_1d(period_vals).flatten()[0])

        if T_target < periods.min() or T_target > periods.max():
            raise ValueError(
                f"目标周期 {T_target:.4f} s 超出可用范围 [{periods.min():.4f}, {periods.max():.4f}] s"
            )

        matches = np.isclose(periods, T_target)
        if matches.any():
            sa_vals, lam_vals = self._read_hazard_curve(files[np.where(matches)[0][0]])
            return np.column_stack([sa_vals, lam_vals])

        hi = np.searchsorted(periods, T_target)
        lo = hi - 1
        T_lo, T_hi = periods[lo], periods[hi]
        sa_lo, lam_lo = self._read_hazard_curve(files[lo])
        sa_hi, lam_hi = self._read_hazard_curve(files[hi])

        sa_common = np.union1d(sa_lo, sa_hi)
        lam_lo_common = np.interp(sa_common, sa_lo, lam_lo)
        lam_hi_common = np.interp(sa_common, sa_hi, lam_hi)

        ln_lam_lo = np.log(lam_lo_common)
        ln_lam_hi = np.log(lam_hi_common)
        weight = (T_target - T_lo) / (T_hi - T_lo)
        ln_lam_interp = ln_lam_lo + weight * (ln_lam_hi - ln_lam_lo)
        lam_interp = np.exp(ln_lam_interp)

        log_sa = np.linspace(np.log(sa_common.min()), np.log(sa_common.max()), resample_points)
        ln_lam_common = np.interp(log_sa, np.log(sa_common), np.log(lam_interp))
        sa_new = np.exp(log_sa)
        lam_new = np.exp(ln_lam_common)
        return np.column_stack([sa_new, lam_new])

    def manual_probability(self,
            EDP_type: str,
            Sa_range: tuple[float, float],
            EDP_range: tuple[float, float],
            hazard_curve: np.ndarray,
            fragility_type: Literal['computed', 'PSDM']='computed',
            density: float=1000,
            *args, **kwargs
        ) -> None | tuple[float, float, float, float]:
        """年度超越概率计算

        Args:
            EDP (str): 工程需求参数类型
            Sa_range (tuple[float, float]): 灾害作用范围
            EDP_range (tuple[float, float]): 工程需求参数范围
            hazard_curve (np.ndarray): 灾害曲线(二维数组，两列)
            fragility_type (Literal['computed', 'PSDM']): 使用哪种易损性曲线('computed': 基于IDA数据直接拟合; 'PSDM': 基于概率地震需求模型)
            density (float): 采样点密度
        """
        internal_call: bool = kwargs.get('internal_call', False)
        density_EDP: int = kwargs.get('density_EDP', density)
        self.hazard_curve = hazard_curve
        Sa_min, Sa_max = Sa_range
        EDP_min, EDP_max = EDP_range
        if EDP_min == 0:
            EDP_min = 1e-8
        x_Sa = np.linspace(Sa_min, Sa_max, density)
        get_log10_harzard_curve = interp1d(
            np.log10(hazard_curve[:, 0]), np.log10(hazard_curve[:, 1]),
            kind='cubic', fill_value='extrapolate', bounds_error=False)
        x_temp = np.linspace(hazard_curve[0, 0], hazard_curve[-1, 0], 1000)

        if EDP_type not in self.EDP_types:
            raise KeyError(f'尚未指定 `{EDP_type}` 类型，请在 `__init__` 的 `EDP_types` 中添加。')

        x_EDP = np.linspace(EDP_min, EDP_max, density_EDP)

        if fragility_type == 'computed':
            y_ls = np.zeros((self.GM_N, density))  # IDA曲线纵坐标(密集点)
            for i, df in enumerate(self.data):
                x_points = [0] + df[EDP_type].to_list()
                y_points = [0] + df['IM'].to_list()
                _, y = get_percentile_line([x_points], [y_points], p=0.5, x=x_EDP)
                y_ls[i] = y
            ln_theta = np.median(np.log(y_ls), axis=0)  # 地震动强度中值
            beta = np.std(np.log(y_ls), axis=0, ddof=1)  # 地震动强度标准差
        gama = np.zeros_like(x_EDP)
        log10_HSa = get_log10_harzard_curve(np.log10(x_Sa))
        HSa = np.power(10, log10_HSa)
        diff_HSa = np.append(0, np.diff(HSa))
        for i in range(len(x_EDP)):
            # 遍历EDP值
            EDP_value = x_EDP[i]
            if fragility_type == 'computed':
                ln_theta_i = ln_theta[i]
                beta_i = beta[i]
                y_cdf = norm.cdf((np.log(x_Sa) - ln_theta_i) / beta_i, 0, 1)  # 易损性函数(累积概率分布曲线)
            elif fragility_type == 'PSDM':
                if not EDP_type in self.EDP_types:
                    raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
                A, B = self.AB[EDP_type]
                y_cdf = norm.cdf((A + B * np.log(x_Sa) - np.log(EDP_value)) / self.beta[EDP_type], 0, 1)
            else:
                raise ValueError(f'Wrong fragility_type: {fragility_type}')
            gama_i = np.sum(y_cdf * np.abs(diff_HSa))  # 易损性函数乘以灾害曲线的差分
            gama[i] = gama_i
        P50years = 1 - np.exp(-gama_i * 50)
        text_risk = f'Risk Evaluation:\nManual probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {gama_i:.3e}\n'
        text_risk += f'50 Years Probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {P50years:.2%}'

        if not internal_call:
            print(text_risk)
            self.risk_text[EDP_type] = text_risk
            self.risk_EDP_hazard_curves[EDP_type] = np.array([x_EDP, gama]).T
            self.has_risk_data = True

            fig,axes = self.all_figures_2[EDP_type]
            ax = axes.flatten()[0]
            ax.loglog(hazard_curve[:, 0], hazard_curve[:, 1], '-o', label='USGS curve')
            ax.loglog(x_temp, pow(10, get_log10_harzard_curve(np.log10(x_temp))), label='Cubic Interpolation')
            ax.grid(True)
            ax.set_xlabel('Sa')
            ax.set_ylabel('MAF of Sa')
            ax.set_title('Hazard Curve')
            ax.legend()

            ax = axes.flatten()[1]
            ax.plot(x_Sa, y_cdf, label=f'P[{EDP_type} > {round(x_EDP[-1], 2)}]')
            ax.set_xlabel('Sa')
            ax.set_ylabel('Probability of Exceedance')
            ax.set_title('Fragility Curve')
            ax.legend()
            ax.grid(True)

            ax = axes.flatten()[2]
            ax.semilogy(x_EDP, gama)
            ax.set_xlabel(EDP_type)
            ax.set_ylabel(f'MAF of {EDP_type}')
            ax.set_title(f'{EDP_type} Hazard Curve')
            ax.grid(True)

        else:
            if np.isnan(beta_i) or np.isnan(ln_theta_i):
                print("[返回警告] manual_probability 返回非法数据，返回 None")
                return None, None, None, None
            return np.exp(ln_theta_i), beta_i, gama_i, P50years


    # def manual_probability(self,
    #         EDP_type: str,
    #         Sa_range: tuple[float, float],
    #         EDP_range: tuple[float, float],
    #         hazard_curve: np.ndarray,
    #         fragility_type: Literal['computed', 'PSDM']='computed',
    #         density: float=1000,
    #         *args, **kwargs
    #     ) -> None | tuple[float, float, float, float]:
    #     """年度超越概率计算

    #     Args:
    #         EDP (str): 工程需求参数类型
    #         Sa_range (tuple[float, float]): 灾害作用范围
    #         EDP_range (tuple[float, float]): 工程需求参数范围
    #         hazard_curve (np.ndarray): 灾害曲线(二维数组，两列)
    #         fragility_type (Literal['computed', 'PSDM']): 使用哪种易损性曲线('computed': 基于IDA数据直接拟合; 'PSDM': 基于概率地震需求模型)
    #         density (float): 采样点密度
    #     """
    #     internal_call: bool = kwargs.get('internal_call', False)
    #     density_EDP: int = kwargs.get('density_EDP', density)
    #     self.hazard_curve = hazard_curve
    #     Sa_min, Sa_max = Sa_range
    #     EDP_min, EDP_max = EDP_range
    #     if EDP_min == 0:
    #         EDP_min = 1e-8
    #     x_Sa = np.linspace(Sa_min, Sa_max, density)  # Sa轴坐标
    #     get_log10_harzard_curve = interp1d(np.log10(hazard_curve[:, 0]), np.log10(hazard_curve[:, 1]), kind='cubic', fill_value='extrapolate', bounds_error=False)
    #     x_temp = np.linspace(hazard_curve[0, 0], hazard_curve[-1, 0], 1000)
    #     if EDP_type not in self.EDP_types:
    #         raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
    #     x_EDP = np.linspace(EDP_min, EDP_max, density_EDP)  # IDA曲线横坐标(密集点)
    #     if fragility_type == 'computed':
    #         y_ls = np.zeros((self.GM_N, density))  # IDA曲线纵坐标(密集点)
    #         for i, df in enumerate(self.data):
    #             x_points = [0] + df[EDP_type].to_list()
    #             y_points = [0] + df['IM'].to_list()
    #             _, y = get_percentile_line([x_points], [y_points], p=0.5, x=x_EDP)
    #             y_ls[i] = y
    #         ln_theta = np.median(np.log(y_ls), axis=0)  # 地震动强度中值
    #         beta = np.std(np.log(y_ls), axis=0, ddof=1)  # 地震动强度标准差
    #     gama = np.zeros_like(x_Sa)
    #     log10_HSa = get_log10_harzard_curve(np.log10(x_Sa))  # 灾害曲线的对数
    #     HSa = np.power(10, log10_HSa)  # 灾害曲线纵坐标
    #     diff_HSa = np.append(0, np.diff(HSa))  # 灾害曲线的差分
    #     for i in range(len(x_EDP)):
    #         # 遍历EDP值
    #         EDP_value = x_EDP[i]
    #         if fragility_type == 'computed':
    #             ln_theta_i = ln_theta[i]
    #             beta_i = beta[i]
    #             y_cdf = norm.cdf((np.log(x_Sa) - ln_theta_i) / beta_i, 0, 1)  # 易损性函数(累积概率分布曲线)
    #         elif fragility_type == 'PSDM':
    #             if not EDP_type in self.EDP_types:
    #                 raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
    #             A, B = self.AB[EDP_type]
    #             y_cdf = norm.cdf((A + B * np.log(x_Sa) - np.log(EDP_value)) / self.beta[EDP_type], 0, 1)
    #         else:
    #             raise ValueError(f'Wrong fragility_type: {fragility_type}')
    #         gama_i = np.sum(y_cdf * np.abs(diff_HSa))  # 易损性函数乘以灾害曲线的差分
    #         gama[i] = gama_i
    #     P50years = 1 - np.exp(-gama_i * 50)
    #     text_risk = f'Risk Evaluation:\nManual probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {gama_i:.3e}\n'
    #     text_risk += f'50 Years Probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {P50years:.2%}'
    #     if not internal_call:
    #         print(text_risk)
    #         self.risk_text[EDP_type] = text_risk
    #         self.risk_EDP_hazard_curves[EDP_type] = np.array([x_EDP, gama]).T
    #         self.has_risk_data = True
    #         # 画图: 灾害曲线
    #         fig = self.all_figures_2[EDP_type]
    #         axes = fig.get_axes()
    #         ax: Axes = axes[0]
    #         ax.loglog(hazard_curve[:, 0], hazard_curve[:, 1], '-o', label='USGS curve')
    #         ax.loglog(x_temp, pow(10, get_log10_harzard_curve(np.log10(x_temp))), label='Cubic Interpolation')
    #         ax.grid(True)
    #         ax.set_xlabel('Sa')
    #         ax.set_ylabel(f'MAF of Sa')
    #         ax.set_title('Hazard Curve')
    #         ax.legend()
    #         ax: Axes = axes[1]
    #         ax.plot(x_Sa, y_cdf, label=f'P[{EDP_type} > {round(x_EDP[-1], 2)}]')
    #         ax.set_xlabel('Sa')
    #         ax.set_ylabel('Probability of Exceedance')
    #         ax.set_title('Fragility Curve')
    #         ax.legend()
    #         ax.grid(True)
    #         ax: Axes = axes[2]
    #         ax.semilogy(x_EDP, gama)
    #         ax.set_xlabel(EDP_type)
    #         ax.set_ylabel(f'MAF of {EDP_type}')
    #         ax.set_title(f'{EDP_type} Hazard Curve')
    #         ax.grid(True)
    #     else:
    #         # 内部调用的情况，用于进行倒塌易损性评估
    #         # 返回倒塌强度中值，标准差，年度超越概率和50年超越概率
    #         return np.exp(ln_theta_i), beta_i, gama_i, P50years

   

    def visualization(self):
        """可视化曲线图"""
        plt.show()

    def collapse_evaluation(self,
            T: float,
            MCE_spec: Path | str,
            miuT: float,
            SF_spec: float=1,
            beta: list[float] | float=None,
            SDC: Literal['B', 'C', 'Dmin', 'Dmax', 'other']= 'other',
        ):
        """Collapse assessment based on FEMA P695 and manual probability

        Args:
            T (float): 结构周期(T=CuTa，非一阶周期)
            MCE_spec (Path | str): MCE规范谱所在文件路径(文件为两列，一列周期一列谱值)
            miuT (float): Period-based ductility
            SF_spec (float, optional): 将规范谱的谱值进行放大，默认放大系数为1
            beta: list[float] | float=None,
            SDC (Literal['B', 'C', 'Dmin', 'Dmax', 'other'], optional): Seismic design category，默认None

        注：
        ------
        不确定性参数`beta`可填多个值，
        多个`beta`值将按下式进行叠加: `beta = sqrt(beta_1^2 + beta_2^2 + beta_3^2 + ...)`
        例如：  
        >>> beta = [0.2, 0.3, 0.4]
        表示beta = sqrt(0.2^2 + 0.3^2 + 0.4^2)
        也可只传入单个不确定值，例如：  
        >>> beta = 0.4  
        """
        self.has_collapse_data = True
        MCE_spec = Path(MCE_spec)
        if not MCE_spec.exists():
            raise FileExistsError(f'无法找到文件：{str(MCE_spec.absolute())}')
        if 'IDR' not in self.EDP_types:
            raise ValueError('未进行`IDR`类型的易损性分析')
        if self.hazard_curve is None:
            raise ValueError('缺少地震危险性数据，请先运行`manual_probability`方法')
        spec_data = np.loadtxt(MCE_spec)  # MCE规范谱
        spec_data[:, 1] *= SF_spec
        S_MT = get_y(spec_data[:, 0], spec_data[:, 1], T)  # 一阶周期对应的MCE谱谱值
        theta, beta_calc, x_collapse_fragility, _, y_collapse_fragility, Sa_collapse, df_clps_frag\
            = self.exceedance_probability('IDR', self.collapse_limit, beta, internal_call=True)
        self.df_clps_frag = df_clps_frag
        # 倒塌强度指标统计
        self.collapse_intensity = {
            'median': theta,  # 中值倒塌强度
            'mean': np.mean(Sa_collapse),  # 平均倒塌强度
            'log_std': beta_calc,  # 倒塌强度的对数标准差
            'std': np.std(Sa_collapse, ddof=1),  # 倒塌强度的标准差
            'SaS_MT_MCE': S_MT,  # 大震规范谱值
            'Sa_collapse': Sa_collapse,  # 倒塌强度曲线对应的Sa值
        }
        # Part-1 基于FEMA P695的倒塌性能评估
        S_CT = get_x(x_collapse_fragility, y_collapse_fragility, 0.5)  # 中值倒塌强度
        Pc_MCE = get_y(x_collapse_fragility, y_collapse_fragility, S_MT)  # MCE下倒塌概率
        CMR = S_CT / S_MT
        ssf = get_SFF(T, miuT, SDC)  # 谱形状修正系数
        ACMR = CMR * ssf
        try:
            iter(beta)
        except TypeError:
            beta = [beta]

        beta_total = 0
        valid_beta = []

        for beta_i in beta:
            if beta_i is None or np.isnan(beta_i):
                logger.warning(f"忽略非法 beta 值: {beta_i}")
                continue
            valid_beta.append(beta_i)
            beta_total += beta_i ** 2

        if valid_beta:
            beta_total = float(np.sqrt(beta_total))
        else:
            logger.error("未提供有效的 beta 值，倒塌评估可能无效。")
            beta_total = 0.0  # 或抛出异常
        ACMR5, ACMR10, ACMR15, ACMR20, ACMR25 = get_acceptable_ACMR(beta_total)
        criteria1 = bool(ACMR >= ACMR10)
        criteria2 = bool(ACMR >= ACMR20)
        criteria3 = bool(Pc_MCE < 0.1)
        self.all_CMR['CMR'] = CMR
        self.all_CMR['ACMR'] = ACMR
        self.all_CMR['ACMR5'] = ACMR5
        self.all_CMR['ACMR10'] = ACMR10
        self.all_CMR['ACMR15'] = ACMR15
        self.all_CMR['ACMR20'] = ACMR20
        self.all_CMR['ACMR25'] = ACMR25
        self.all_CMR['beta_TOT'] = beta_total
        self.all_CMR['Pc_MCE'] = Pc_MCE
        self.all_CMR['ssf'] = ssf
        self.all_CMR['ACMR>ACMR10'] = criteria1
        self.all_CMR['ACMR>ACMR20'] = criteria2
        if all([criteria1, criteria2, criteria3]):
            logger.success('倒塌性能满足FEMA P695要求')
        else:
            if not criteria1:
                logger.warning(f'倒塌性能不满足FEMA P695要求: ACMR < ACMR10 ({ACMR:.3f} < {ACMR10:.3f})')
            if not criteria2:
                logger.warning(f'倒塌性能不满足FEMA P695要求: ACMR < ACMR20 ({ACMR:.3f} < {ACMR20:.3f})')
            if not criteria3:
                logger.warning(f'倒塌性能不满足FEMA P695要求: Pc_MCE > 10 % ({Pc_MCE:.2%} > 10 %)')
        print('Collapse assessment based on FEMA P695:')
        print(f'computed beta = {beta_calc:.3f} (对数标准差)')
        print(f'S_CT = {S_CT:.3f} (中值倒塌强度), S_MT = {S_MT:.3f} (大震规范谱值)')
        print(f'Pc_MCE = {Pc_MCE:.2%}')
        print(f'CMR = {CMR:.3f}')
        print(f'SSF = {ssf:.3f}')
        print(f'ACMR = {ACMR:.3f}')
        print(f'ACMR10 = {ACMR10:.3f}')
        print(f'ACMR20 = {ACMR20:.3f}')
        # Part-2 基于风险的倒塌评估
        hazard_curve = self.hazard_curve
        Sa_range = (hazard_curve[0, 0], hazard_curve[-1, 0])
        EDP_range = (self.collapse_limit, self.collapse_limit)
        fragility_type = 'computed'
        density = 1000
        median, std, gama, P50years = self.manual_probability('IDR', Sa_range, EDP_range, hazard_curve, fragility_type, density, internal_call=True, density_EDP=1)
        print('Collapse assessment based on seismic risk:')
        print(f'P(1 year) = {gama:.3e}')
        print(f'P(50 years) = {P50years:.2%}')
        self.collapse_risk['gama'] = gama
        self.collapse_risk['P50'] = P50years

    def save_data(self, output_path: str | Path):
        """保存分析结果
            
            Args:
                output_path (str | Path): 保存结果的文件夹路径
        """
        output_path = Path(output_path)
        if not Path.exists(output_path):
            os.makedirs(output_path)
        print('正在保存数据...\r', end='')
        for EDP_type in self.EDP_types:
            if EDP_type in self.all_figures_1:
                fig, _ = self.all_figures_1[EDP_type]
                fig.savefig(output_path / f'FragilityAnalysis_{EDP_type}.png', dpi=600)
            if EDP_type in self.all_figures_2:
                fig, _ = self.all_figures_2[EDP_type]
                fig.savefig(output_path / f'RiskAnalysis_{EDP_type}.png', dpi=600)
            # 1 IDA曲线
            wb = px.Workbook()
            ws1 = wb.active
            ws1.title = 'IDA曲线'
            for i, (x, y) in enumerate(zip(self.DM_lines[EDP_type], self.IM_lines[EDP_type])):
                ws1.cell(1, 2*i+1, i+1)
                ws1.merge_cells(start_row=1, start_column=2*i+1, end_row=1, end_column=2*i+2)
                for j, (xi, yi) in enumerate(zip(x, y)):
                    ws1.cell(j+2, 2*i+1, value=xi)
                    ws1.cell(j+2, 2*i+2, value=yi)
            ws2 = wb.create_sheet('分位线')
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws2.cell(1, 1, '16%分位线')
            ws2.cell(1, 3, '50%分位线')
            ws2.cell(1, 5, '84%分位线')
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_16[EDP_type])):
                ws2.cell(i+2, 1, x)
                ws2.cell(i+2, 2, y)
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_50[EDP_type])):
                ws2.cell(i+2, 3, x)
                ws2.cell(i+2, 4, y)
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_84[EDP_type])):
                ws2.cell(i+2, 5, x)
                ws2.cell(i+2, 6, y)
            wb.save(output_path / f'IDA曲线_{EDP_type}.xlsx')
            # 2 概率需求模型
            wb = px.Workbook()
            ws = wb.active
            ws.title = 'ln(DM)-ln(IM)'
            ws.cell(1, 1, 'ln(DM)-ln(IM)散点')
            ws.cell(1, 3, '拟合值')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
            if EDP_type in self.ln_IM_line.keys():
                for i, (x, y) in enumerate(zip(self.ln_IM[EDP_type], self.ln_DM[EDP_type])):
                    ws.cell(i+2, 1, x)
                    ws.cell(i+2, 2, y)
                for i, (x, y) in enumerate(zip(self.ln_IM_line[EDP_type], self.ln_DM_line[EDP_type])):
                    ws.cell(i+2, 3, x)
                    ws.cell(i+2, 4, y)
            wb.save(output_path / f'概率需求模型_{EDP_type}.xlsx')
            # 3 地震易损性、超越概率(倒塌易损性)
            wb = px.Workbook()
            ws1 = wb.active
            ws1.title = '地震易损性'
            ws1.cell(1, 1, 'IM')
            ws1.cell(1, 2, '超越概率')
            if EDP_type in self.x_frag.keys():
                ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
                ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+len(self.DS[EDP_type]))
                for i, label in enumerate(self.DS[EDP_type].keys()):
                    ws1.cell(2, i+2, label)
                for i, x in enumerate(self.x_frag[EDP_type]):
                    ws1.cell(3+i, 1, x)
                for i in range(len(self.y_frag[EDP_type])):
                    for j, y in enumerate(self.y_frag[EDP_type][i]):
                        ws1.cell(3+j, i+2, y)
            if EDP_type in self.exceed_x and self.exceed_x[EDP_type]:
                for label, DM_value in self.DM_values.get(EDP_type, {}).items():
                    ws2 = wb.create_sheet(f'超越概率P({EDP_type}>{DM_value})')
                    ws2.cell(1, 1, '实际(散点)')
                    ws2.cell(1, 3, '拟合(曲线)')
                    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                    ws2.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
                    for i, (x, y) in enumerate(zip(self.exceed_x[EDP_type][label], self.exceed_y[EDP_type][label])):
                        ws2.cell(2+i, 1, x)
                        ws2.cell(2+i, 2, y)
                    for i, (x, y) in enumerate(zip(self.exceed_x_fit[EDP_type][label], self.exceed_y_fit[EDP_type][label])):
                        ws2.cell(2+i, 3, x)
                        ws2.cell(2+i, 4, y)
                    if label in self.exceed_y_fixedBeta.get(EDP_type, {}):
                        ws2.cell(1, 5, '拟合(固定beta_TOT)')
                        ws2.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
                        for i, (x, y) in enumerate(zip(self.exceed_x_fit[EDP_type][label], self.exceed_y_fixedBeta[EDP_type][label])):
                            ws2.cell(2+i, 5, x)
                            ws2.cell(2+i, 6, y)
            wb.save(output_path / f'易损性曲线_{EDP_type}.xlsx')
            # 保存计算结果参数
            if EDP_type in self.info.keys():
                with open(output_path / f'概率特征_{EDP_type}.out', 'w') as f:
                    f.write(self.info[EDP_type])
        # 保存风险评估结果
        if self.has_risk_data:
            for EDP_type in self.risk_EDP_hazard_curves.keys():
                with open(output_path / f'risk_info_{EDP_type}.out', 'w') as f:
                    f.write(self.risk_text[EDP_type])
                np.savetxt(output_path / f'hazard_curve_{EDP_type}.out', self.risk_EDP_hazard_curves[EDP_type], fmt='%.4e')
        # 保存倒塌评估结果
        if self.has_collapse_data:
            with open(output_path / f'CMR.json', 'w') as f:
                json.dump(self.all_CMR, f, indent=4)
            with open(output_path / f'collapse_risk.json', 'w') as f:
                json.dump(self.collapse_risk, f, indent=4)
            with open(output_path / f'collapse_intensity.json', 'w') as f:
                json.dump(self.collapse_intensity, f, indent=4)
            self.df_clps_frag.to_csv(output_path / f'倒塌易损性曲线.csv', index=False)
        logger.success('已保存数据')

        def __del__(self):
            for _, (fig, _) in self.all_figures_1.items():
                fig.clear()
            for _, (fig, _) in self.all_figures_2.items():
                fig.clear()
            plt.cla()
            plt.clf()
            plt.close('all')



if __name__ == "__main__":

    i = 'SMAPFDF'
    model = FragilityAnalysis(
        f'D:\Study\My_Project\OpenSAS-SMAPFDB\Output_data\MC8_{i}\MC8_IDA_data_out',
        EDP_types=['IDR', 'RIDR','PFA'],
        collapse_limit=0.1,
       )
    model.calc_IDA('IDR')
    model.calc_IDA('RIDR')
    model.calc_IDA('PFA')
    model.frag_curve('IDR', {'DS-1': 0.005, 'DS-2': 0.01, 'DS-3': 0.02, 'DS-4': 0.04}, beta=0.4)
    model.frag_curve('RIDR', {'DS-1': 0.001, 'DS-2': 0.002, 'DS-3': 0.005, 'DS-4': 0.01}, beta=0.4)
    model.frag_curve('PFA', {'DS-1': 0.2, 'DS-2': 0.5, 'DS-3': 1}, beta=0.4)
    model.exceedance_probability('IDR', 0.1)
    model.collapse_evaluation(T1=2.613, MCE_spec=r'D:\Study\My_Project\OpenSAS-SMAPFDB\Spectrum\MCE Level Spectrum.txt', SF_spec=1)
    model.visualization()
    # model.save_data(f'D:\Study\My_Project\OpenSAS-SMAPFDB\Output_data\MC8_{i}\MC8_IDA_data_frag')

    
